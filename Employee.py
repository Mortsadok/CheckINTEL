import ast
import DateTime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import urllib.request
from threading import Thread


#  Init Remembered Values:

# {'11957404': '3', '11957403': '3', '11957402': '3', '11951196': '3', '11721283': '3', '11718878': '3', '11763904': '3',
# '11840167': '3', '11987649': '3', '11969967': '3', '11964678': '3', '12021975': '3', '11588566': '3', '11888572': '3',
# '11424368': '3', '11949749': '3', '11400556': '3', '11773043': '3', '11656032': '3', '12011864': '3'}

#  Init Employee State:

# {'11957404': '0', '11957403': '0', '11957402': '0', '11951196': '0', '11721283': '0', '11718878': '0', '11763904': '0',
# '11840167': '0', '11987649': '0', '11969967': '0', '11964678': '0', '12021975': '0', '11588566': '0', '11888572': '0',
# '11424368': '0', '11949749': '0', '11400556': '0', '11773043': '0', '11656032': '0', '12011864': '0'}


def AdminSMS(Employee, TeamLead):
    MorPhone = '0502241114'
    OmerPhone = '0549876885'

    if TeamLead == 'Mor':
        Message = f'Please%20Note,%0A{Employee}`s%20Check-In%20time:%0A{DateTime.Date()}%20{DateTime.Time()[:5]}'
        url = f"http://smsdirect.cellcom.co.il/SmsGate/SmsGate2.asmx/SendSms?Username=intelhif&Password=hif5047&Target={MorPhone}" \
              f"&Source=CheckAdmin&Message={Message}&Validity=0&Replace=false&Immediate=false"
        urllib.request.urlopen(url).read()

    if TeamLead == 'Omer':
        Message = f'Please%20Note,%0A{Employee}`s%20Check-In%20time:%0A{DateTime.Date()}%20{DateTime.Time()[:5]}'
        url = f"http://smsdirect.cellcom.co.il/SmsGate/SmsGate2.asmx/SendSms?Username=intelhif&Password=hif5047&Target={OmerPhone}" \
              f"&Source=CheckAdmin&Message={Message}&Validity=0&Replace=false&Immediate=false"
        urllib.request.urlopen(url).read()


def SMS(Employee, Phone, Check):
    if Check == 1:
        try:
            Message = f'Welcome%20to%20Intel,%20{Employee}!%0ACheck-In%20time:%20{DateTime.Date()}%20{DateTime.Time()[:5]}'
            url = f"http://smsdirect.cellcom.co.il/SmsGate/SmsGate2.asmx/SendSms?Username=intelhif&Password=hif5047&Target={Phone}" \
                  f"&Source=Quali-Check&Message={Message}&Validity=0&Replace=false&Immediate=false"
            urllib.request.urlopen(url).read()
        except IOError:
            pass

    if Check == 0:
        try:
            Message = f'Good%20bye,%20{Employee}!%0ACheck-Out%20time:%20{DateTime.Date()}%20{DateTime.Time()[:5]}'
            url = f"http://smsdirect.cellcom.co.il/SmsGate/SmsGate2.asmx/SendSms?Username=intelhif&Password=hif5047&Target={Phone}" \
                  f"&Source=Quali-Check&Message={Message}&Validity=0&Replace=false&Immediate=false"
            urllib.request.urlopen(url).read()
        except IOError:
            pass


def MakeAnObject():
    ALLEMP = []
    EMP_Database = open(r'Dependencies\EmployeeDatabase.txt', "r").read()
    Employees = EMP_Database.split('\n')

    for EMP in Employees:
        currentEMP = EMP.split(' ')
        E1 = Employee(currentEMP[0], currentEMP[1], currentEMP[2], currentEMP[3], currentEMP[4])
        ALLEMP.append(E1)

    return ALLEMP


def EmployeeCheck(Values):
    Value = open(r'Dependencies\EmployeeCheck.txt', "w")
    Value.write(str(Values))
    Value.close()
    return Value


def EmployeeState(Values):
    Value = open(r'Dependencies\EmployeeState.txt', "w")
    Value.write(str(Values))
    Value.close()
    return Value


def Value_Is_Changed(Values):
    Value = open(r'Dependencies\Remembered_Values.txt', "w")
    Value.write(str(Values))
    Value.close()
    return Value


def toExcel(Excel, Manager):  # Employee, Lastname, State, Location
    if Manager == 'Mor':
        workbook = load_workbook('Dependencies\SV.xlsx')

    else:
        workbook = load_workbook('Dependencies\LAB.xlsx')

    for WriteToExcel in Excel:
        Employee = WriteToExcel[0]
        Lastname = WriteToExcel[1]
        State = WriteToExcel[2]
        Location = WriteToExcel[3]

        if Employee == 'Mor':
            if Lastname == 'Tsadok':
                workbook.active = 0
            if Lastname == 'Shilo':
                workbook.active = 6

        else:
            for ESheet in range(0, len(workbook.sheetnames)):
                if workbook.sheetnames[ESheet] == Employee:
                    workbook.active = ESheet

        sheet = workbook.active
        sheet[Location] = State
        sheet.alignment = Alignment(vertical='center')

    if Manager == 'Mor':
        workbook.save(filename='Dependencies\SV.xlsx')
    else:
        workbook.save(filename='Dependencies\LAB.xlsx')


class Employee:

    def __init__(self, Firstname, Lastname, WWID, Phone, TeamLead, CheckOut=None, CheckIn=None, EndOfTheDay=0):
        self.Firstname = Firstname
        self.Lastname = Lastname
        self.WWID = WWID
        self.Phone = Phone
        self.TeamLead = TeamLead
        self.CheckOut = CheckOut
        self.CheckIn = CheckIn
        self.EndOfTheDay = EndOfTheDay

    def WWID_SCAN(self, Location, Number, Queue):
        try:
            State = open(r'Dependencies\EmployeeState.txt', "r").read()
            EState = ast.literal_eval(State)

            Check = open(r'Dependencies\EmployeeCheck.txt', "r").read()
            CState = ast.literal_eval(Check)

            if EState[self.WWID] == '1':

                if CState[self.WWID] != DateTime.Date():

                    Location = Location.replace('E', 'F')
                    self.CheckOut = 'Missing Check-Out!'
                    toSend = [[self.Firstname, self.Lastname, self.CheckOut, Location]]
                    toExcel(toSend, self.TeamLead)
                    Location = Location.replace('F', 'E')
                    Location = Location[0] + str(int(Number) + 1)
                    Values = open(r'Dependencies\Remembered_Values.txt', "r").read()
                    ENUM = ast.literal_eval(Values)
                    ENUM[self.WWID] = str(int(ENUM[self.WWID]) + 1)
                    Value_Is_Changed(ENUM)

                    self.CheckOut = None
                    self.CheckIn = None
                    self.EndOfTheDay = 1
                    EState[self.WWID] = '0'
                    EmployeeState(EState)

                else:
                    Attached = []
                    Location = Location.replace('E', 'F')
                    self.CheckOut = DateTime.Time()[:5]
                    to = [self.Firstname, self.Lastname, self.CheckOut, Location]
                    Attached.append(to)
                    Location = Location.replace('F', 'G')
                    toAttach = [self.Firstname, self.Lastname, 'Working Day', Location]
                    Attached.append(toAttach)
                    toExcel(Attached, self.TeamLead)

                    self.EndOfTheDay = 1
                    Thread(target=SMS, args=(self.Firstname, self.Phone, 0,)).start()
                    EState[self.WWID] = '0'
                    EmployeeState(EState)
                    Q = False
                    Queue.put(Q)
                    return Queue

            if EState[self.WWID] == '0':
                self.CheckIn = DateTime.Time()[:5]
                if len(Location) < 3:
                    Attached = [
                        [self.Firstname, self.Lastname, self.Firstname + " " + self.Lastname, 'D' + Location[1]],
                        [self.Firstname, self.Lastname, DateTime.Date(), 'B' + Location[1]],
                        [self.Firstname, self.Lastname, DateTime.Day(), 'C' + Location[1]],
                        [self.Firstname, self.Lastname, self.CheckIn, Location]]
                    toExcel(Attached, self.TeamLead)
                else:
                    Attached = [[self.Firstname, self.Lastname, self.Firstname + " " + self.Lastname,
                                 'D' + Location[1] + Location[2]],
                                [self.Firstname, self.Lastname, DateTime.Date(), 'B' + Location[1] + Location[2]],
                                [self.Firstname, self.Lastname, DateTime.Day(), 'C' + Location[1] + Location[2]],
                                [self.Firstname, self.Lastname, self.CheckIn, Location]]
                    toExcel(Attached, self.TeamLead)

                if (int(DateTime.Time()[0])) > 0:
                    Thread(target=AdminSMS, args=(self.Firstname, self.TeamLead,)).start()

                self.EndOfTheDay = 0
                Thread(target=SMS, args=(self.Firstname, self.Phone, 1,)).start()
                EState[self.WWID] = '1'
                EmployeeState(EState)
                CState[self.WWID] = DateTime.Date()
                EmployeeCheck(CState)
                Q = True
                Queue.put(Q)
                return Queue

        except Exception as e:
            f = open("Logs/Errors.txt", "a")
            f.write(f'Employee: {self.Firstname + " " + self.Lastname}\n'
                    f'WWID: {self.WWID}\n'
                    f'Date: {DateTime.Date()} {DateTime.Time()}\n'
                    f'Error: {e}')
            f.write('\n\n')
            f.close()


Employees = MakeAnObject()
