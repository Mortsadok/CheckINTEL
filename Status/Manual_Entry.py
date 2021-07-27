import ast
import DateTime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys


def toExcel(Excel, Manager):  # Employee, Lastname, State, Location
    if Manager == 'Mor':
        workbook = load_workbook('../Dependencies/SV.xlsx')

    else:
        workbook = load_workbook('../Dependencies/LAB.xlsx')

    for WriteToExcel in Excel:
        Employee = WriteToExcel[0]
        Lastname = WriteToExcel[1]
        State = WriteToExcel[2]
        Location = WriteToExcel[3]

        if Employee == 'Mor':
            if Lastname == 'Tsadok':
                workbook.active = 0
            if Lastname == 'Shilo':
                workbook.active = 6

        else:
            for ESheet in range(0, len(workbook.sheetnames)):
                if workbook.sheetnames[ESheet] == Employee:
                    workbook.active = ESheet

        sheet = workbook.active
        sheet[Location] = State
        sheet.alignment = Alignment(vertical='center')

    if Manager == 'Mor':
        workbook.save(filename='../Dependencies/SV.xlsx')
    else:
        workbook.save(filename='../Dependencies/LAB.xlsx')


def Value_Is_Changed(Values):
    Value = open(r'../Dependencies/Remembered_Values.txt', "w")
    Value.write(str(Values))
    Value.close()
    return Value


def EmployeeState(Values):
    Value = open(r'../Dependencies/EmployeeState.txt', "w")
    Value.write(str(Values))
    Value.close()
    return Value


def EmployeeCheck(Values):
    Value = open(r'../Dependencies/EmployeeCheck.txt', "w")
    Value.write(str(Values))
    Value.close()
    return Value


def MakeAnObject():
    ALLEMP = []
    EMP_Database = open(r'../Dependencies/EmployeeDatabase.txt', "r").read()
    Employees = EMP_Database.split('\n')

    for EMP in Employees:
        currentEMP = EMP.split(' ')
        E1 = Employee(currentEMP[0], currentEMP[1], currentEMP[2], currentEMP[4])
        ALLEMP.append(E1)

    return ALLEMP


class Employee:

    def __init__(self, Firstname, Lastname, WWID, TeamLead, CheckOut=None, CheckIn=None):
        self.Firstname = Firstname
        self.Lastname = Lastname
        self.WWID = WWID
        self.TeamLead = TeamLead
        self.CheckOut = CheckOut
        self.CheckIn = CheckIn

    def WWID_SCAN(self, Location, Entry):
        toAttach = []

        if len(Location) < 3:
            Attached = [[self.Firstname, self.Lastname, self.Firstname + " " + self.Lastname, 'D' + Location[1]],
                        [self.Firstname, self.Lastname, DateTime.Date(), 'B' + Location[1]],
                        [self.Firstname, self.Lastname, DateTime.Day(), 'C' + Location[1]]]
            toExcel(Attached, self.TeamLead)
        else:
            Attached = [
                [self.Firstname, self.Lastname, self.Firstname + " " + self.Lastname, 'D' + Location[1] + Location[2]],
                [self.Firstname, self.Lastname, DateTime.Date(), 'B' + Location[1] + Location[2]],
                [self.Firstname, self.Lastname, DateTime.Day(), 'C' + Location[1] + Location[2]]]
            toExcel(Attached, self.TeamLead)

        toAttach.append([self.Firstname, self.Lastname, Entry, Location])

        toExcel(toAttach, self.TeamLead)


def Status(WWID, Status):
    Check = open(r'../Dependencies/EmployeeCheck.txt', "r").read()
    CState = ast.literal_eval(Check)

    Values = open(r'../Dependencies/Remembered_Values.txt', "r").read()
    ENUM = ast.literal_eval(Values)

    State = open(r'../Dependencies/EmployeeState.txt', "r").read()
    EState = ast.literal_eval(State)

    Let = 'E'

    for Employee in Employees:
        if WWID == Employee.WWID:
            if EState[Employee.WWID] == '1':
                ENUM[Employee.WWID] = str(int(ENUM[Employee.WWID]) + 1)
                Value_Is_Changed(ENUM)

            Location = Let + str(ENUM.get(WWID))
            Employee.WWID_SCAN(Location, Status)
            EState[Employee.WWID] = '1'
            EmployeeState(EState)

            CState[WWID] = DateTime.Date()
            EmployeeCheck(CState)


Employees = MakeAnObject()
Status(sys.argv[1], sys.argv[2])
